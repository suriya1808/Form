from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime
import json

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# Excel file path - use environment variable for cloud deployment
EXCEL_FILE = os.getenv('EXCEL_FILE_PATH', 'responses.xlsx')

def initialize_excel_file():
    """Initialize the Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "AI Solutions Requests"
        
        # Define headers
        headers = [
            'Submission Time',
            'UserName', 
            'Email',
            'Employee ID',
            'Tower',
            'Problem',
            'Business Benefit',
            'Justification/UseCase'
        ]
        
        # Add headers to the first row
        for col, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col, value=header)
        
        # Apply styling to headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        workbook.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def save_to_excel(data):
    """Save form data to Excel file"""
    try:
        # Load existing workbook
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        
        # Find the next empty row
        next_row = worksheet.max_row + 1
        
        # Prepare data in the correct order
        row_data = [
            data.get('submissionTime', datetime.now().isoformat()),
            data.get('userName', ''),
            data.get('email', ''),
            data.get('employeeId', ''),
            data.get('tower', ''),
            data.get('problem', ''),
            data.get('businessBenefit', ''),
            data.get('justification', '')
        ]
        
        # Add data to the new row
        for col, value in enumerate(row_data, 1):
            worksheet.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        workbook.save(EXCEL_FILE)
        
        print(f"Successfully saved submission to row {next_row}")
        return True
        
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        return False

# Static file routes for serving the web application
@app.route('/')
def serve_index():
    """Serve the main HTML file"""
    return send_from_directory('.', 'index.html')

@app.route('/style.css')
def serve_css():
    """Serve the CSS file"""
    return send_from_directory('.', 'style.css')

@app.route('/script.js')
def serve_js():
    """Serve the JavaScript file"""
    return send_from_directory('.', 'script.js')

@app.route('/health')
def health_check():
    """Health check endpoint for monitoring"""
    return jsonify({
        'status': 'healthy',
        'message': 'AI Solutions Request Flask Server is running!',
        'excel_file': EXCEL_FILE,
        'file_exists': os.path.exists(EXCEL_FILE)
    })

@app.route('/api/status')
def api_status():
    """API status route to check if server is running"""
    return jsonify({
        'message': 'AI Solutions Request Flask Server is running!',
        'status': 'active',
        'excel_file': EXCEL_FILE,
        'file_exists': os.path.exists(EXCEL_FILE)
    })

@app.route('/submit', methods=['POST'])
def submit_survey():
    """Handle form submission"""
    try:
        # Get JSON data from request
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data received'}), 400
        
        # Validate required fields
        required_fields = [
            'userName', 'email', 'employeeId', 'tower', 'problem', 
            'businessBenefit', 'justification'
        ]
        
        missing_fields = []
        for field in required_fields:
            if field not in data or not str(data[field]).strip():
                missing_fields.append(field)
        
        if missing_fields:
            return jsonify({
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }), 400
        
        # Validate email format (basic validation)
        email = data.get('email', '')
        if '@' not in email or '.' not in email:
            return jsonify({'error': 'Invalid email format'}), 400
        
        # Validate tower selection
        valid_towers = ['Tz', 'EDL', 'ESB', 'EDI', 'Corporate', 'Legacy Apps']
        if data.get('tower') not in valid_towers:
            return jsonify({'error': 'Invalid tower selection'}), 400
        
        # Validate problem selection
        valid_problems = [
            'Incident Solution Recommendation', 'Knowledge chatbot', 'Ticket Classification',
            'Dynamic Query Response', 'Automated Ticket Auditing', 'Code Documentation',
            'Sentiment Analysis', 'File Processor for Special Character Corrections'
        ]
        if data.get('problem') not in valid_problems:
            return jsonify({'error': 'Invalid problem selection'}), 400
        
        # Save to Excel
        if save_to_excel(data):
            return jsonify({
                'message': 'AI Solution request submitted successfully!',
                'status': 'success',
                'submission_time': data.get('submissionTime', datetime.now().isoformat())
            }), 200
        else:
            return jsonify({'error': 'Failed to save data to Excel file'}), 500
            
    except Exception as e:
        print(f"Error processing submission: {str(e)}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/stats')
def get_stats():
    """Get statistics about submissions"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({
                'total_submissions': 0,
                'file_exists': False,
                'message': 'No submissions yet'
            })
        
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        
        # Count submissions (excluding header row)
        total_submissions = worksheet.max_row - 1 if worksheet.max_row > 1 else 0
        
        return jsonify({
            'total_submissions': total_submissions,
            'file_exists': True,
            'excel_file': EXCEL_FILE,
            'last_updated': datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE)).isoformat()
        })
        
    except Exception as e:
        return jsonify({'error': f'Error getting stats: {str(e)}'}), 500

@app.route('/download-excel')
def download_excel():
    """Download the Excel file"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'error': 'No Excel file found'}), 404
        
        return send_from_directory('.', EXCEL_FILE, as_attachment=True, 
                                 download_name=f'ai_solutions_responses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    except Exception as e:
        return jsonify({'error': f'Error downloading file: {str(e)}'}), 500

@app.route('/view-data')
def view_data():
    """View all submissions in JSON format"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return jsonify({'error': 'No data found', 'submissions': []})
        
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        worksheet = workbook.active
        
        # Get headers
        headers = [cell.value for cell in worksheet[1]]
        
        # Get all data rows
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                row_dict = dict(zip(headers, row))
                data.append(row_dict)
        
        return jsonify({
            'total_submissions': len(data),
            'submissions': data,
            'excel_file': EXCEL_FILE,
            'last_updated': datetime.fromtimestamp(os.path.getmtime(EXCEL_FILE)).isoformat()
        })
        
    except Exception as e:
        return jsonify({'error': f'Error reading data: {str(e)}'}), 500

@app.errorhandler(404)
def not_found(error):
    return jsonify({'error': 'Endpoint not found'}), 404

@app.errorhandler(500)
def internal_error(error):
    return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':
    # Initialize Excel file on startup
    initialize_excel_file()
    
    print("=" * 50)
    print("AI Solutions Request Flask Server")
    print("=" * 50)
    print(f"Excel file: {EXCEL_FILE}")
    
    # Get port from environment variable (Render uses this)
    port = int(os.environ.get('PORT', 5000))
    print(f"Server starting on port {port}")
    
    print("Available endpoints:")
    print("  GET  /          - Server status")
    print("  POST /submit    - Submit AI solution request")
    print("  GET  /stats     - View statistics")
    print("  GET  /download-excel - Download Excel file")
    print("  GET  /view-data  - View data in JSON format")
    print("=" * 50)
    
    # Run the Flask app with production settings for Render
    app.run(debug=False, host='0.0.0.0', port=port)
